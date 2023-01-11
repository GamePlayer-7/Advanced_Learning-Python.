from openpyxl import load_workbook

wb = load_workbook('wb1.xlsx')

sheet = wb.active

print(f"Max row in the active sheet is {sheet.max_row}")
print(f"Max column in the active sheet is {sheet.max_column}")
