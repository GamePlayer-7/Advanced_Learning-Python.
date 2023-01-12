from openpyxl import load_workbook

wb = load_workbook('wb1.xlsx')

sheet = wb.active

# Access all cells from between row1 & row 2, column 1 & column 3.
for row in sheet.iter_rows(min_row=1,
                           max_row=2,
                           min_col=1,
                           max_col=3):
  
    print([data.value for data in row])
