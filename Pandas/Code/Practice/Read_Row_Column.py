from openpyxl import load_workbook

wb = load_workbook('wb1.xlsx')

sheet = wb.active

print(sheet.cell(row=1, column=1).value)
print(sheet.cell(row=1, column=2).value)
print(sheet.cell(row=1, column=3).value)
print(sheet.cell(row=1, column=4).value)
